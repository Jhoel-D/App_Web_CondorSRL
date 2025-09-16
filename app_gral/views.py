# ----- Módulos principales de Django -----
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.forms import modelformset_factory
from django.utils.timezone import now, localtime
from django.db import transaction
from django.core.paginator import Paginator

# ----- Modelos de la aplicación -----
from .models import Usuario, ProductoInventario, Categoria, Ventas, ProductosVenta, Pedidos, ProductosPedido, Cliente

# ----- Autenticación y usuarios -----
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm, PasswordChangeForm
from django.contrib.auth.models import Group
from django.contrib.auth import update_session_auth_hash

# ----- ORM y consultas en modelos -----
from django.db.models import Sum, Count, Q   # Q = OR en queries
from django.db.models.functions import TruncDate

# ----- Formularios personalizados -----
from .forms import (
    UsuarioUpdateForm, UsuarioForm, UsuarioCreateForm,
    ProductoInventarioForm, ProductoInventarioCreateForm,
    CategoriaForm, CategoriaCreateForm,
    ProductosVentaForm, VentasForm, PedidosForm,
    ClienteForm, ClienteCreateForm
)

# ----- Manejo de mensajes en vistas -----
from django.contrib import messages

# ----- Manejo de fechas y tiempos -----
from datetime import datetime

# ----- Reportes en Excel (openpyxl) -----
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from io import BytesIO

# ----- Reportes en PDF (ReportLab) -----
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch


# Menú Principal
def menu_principal(request):
    es_administrador = request.user.groups.filter(name='Administrador').exists()
    return render(request, 'menu_principal.html', {'es_administrador': es_administrador})

# Autenticación
def login_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("login")
        return view_func(request, *args, **kwargs)  # ejecuta la vista original
    return wrapper

def iniciar_sesion(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)  # Crea una instancia del formulario con los datos del POST
        if form.is_valid():  # Verifica si el formulario es válido
            user = form.get_user()  # Obtiene el usuario del formulario
            login(request, user)  # Inicia sesión al usuario
            return redirect('menu_principal')  # Redirige a la vista de cómics y mangas
        else:
            return render(request, 'iniciar_sesion.html', {
                'form': form,
                'error_message': 'Usuario o contraseña incorrectos'
            })
    else:
        form = AuthenticationForm()  # Crea una instancia vacía del formulario
    return render(request, 'iniciar_sesion.html', {'form': form})  # Renderiza la plantilla de inicio de sesión

@login_required
def cerrar_sesion (request):
    logout(request)
    return redirect ('menu_principal')  
#MOD USUARIOS
@login_required
def mod_usuarios_home(request):
    # Obtener usuarios
    todos_usuarios = Usuario.objects.all()

    # Separar usuarios activos e inactivos
    usuarios_activos = todos_usuarios.filter(is_active=True)
    usuarios_inactivos = todos_usuarios.filter(is_active=False)

    # Filtro de búsqueda
    search_term = request.GET.get('search', '')
    if search_term:
        usuarios_activos = usuarios_activos.filter(
            Q(first_name__icontains=search_term) |
            Q(last_name__icontains=search_term) |
            Q(CI__icontains=search_term) |
            Q(username__icontains=search_term)
        )
        usuarios_inactivos = usuarios_inactivos.filter(
            Q(first_name__icontains=search_term) |
            Q(last_name__icontains=search_term) |
            Q(CI__icontains=search_term) |
            Q(username__icontains=search_term)
        )

    # Paginación para usuarios activos
    paginator_activos = Paginator(usuarios_activos, 8)  # 10 usuarios activos por página
    page_number_activos = request.GET.get('page_activos')
    usuarios_activos_page = paginator_activos.get_page(page_number_activos)

    # Paginación para usuarios inactivos
    paginator_inactivos = Paginator(usuarios_inactivos, 8)  # 10 usuarios inactivos por página
    page_number_inactivos = request.GET.get('page_inactivos')
    usuarios_inactivos_page = paginator_inactivos.get_page(page_number_inactivos)

    # Renderizar la plantilla con usuarios paginados
    return render(request, 'usuarios/mod_personal/mod_usuarios_home.html', {
        'usuarios_activos': usuarios_activos_page,
        'usuarios_inactivos': usuarios_inactivos_page,
        'search_term': search_term
    })

@login_required
def ver_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    return render(request, 'usuarios/mod_personal/ver_usuario.html', {'usuario': usuario})

@login_required
def editar_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)

    if request.method == 'POST':
        form = UsuarioForm(request.POST, request.FILES, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario editado exitosamente.')
            return redirect('mod_usuarios_home')
        else:
            # Imprimir errores en la consola para depuración
            print(form.errors)
            messages.error(request, 'Por favor, corrige los errores del formulario.')
    else:
        form = UsuarioForm(instance=usuario)

    return render(request, 'usuarios/mod_personal/editar_usuario.html', {'form': form, 'usuario': usuario})

@login_required
def eliminar_usuario(request, usuario_id):
    # usuario = get_object_or_404(Usuario, id=usuario_id)
    # usuario.delete()
    # messages.success(request, 'Usuario eliminado exitosamente.')
    # return redirect('mod_usuarios_home')
     # Obtener el usuario o retornar 404 si no se encuentra
    usuario = get_object_or_404(Usuario, id=usuario_id)

    # Cambiar el estado del usuario
    if usuario.is_active:
        usuario.is_active = False  # Desactivar el usuario (Dado de baja)
        estado = "Dado de baja"
    else:
        usuario.is_active = True   # Activar el usuario
        estado = "Activo"

    usuario.save()  # Guardar los cambios en el estado

    # Mostrar un mensaje de éxito
    messages.success(request, f'Usuario {estado} exitosamente.')

    # Redirigir a la vista de la lista de usuarios
    return redirect('mod_usuarios_home')

@login_required
def registrar_usuario(request):
    if request.method == 'POST':
        form = UsuarioCreateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()  # Esto ya maneja la asociación del grupo en el método save()
            messages.success(request, 'Usuario creado exitosamente.')
            return redirect('mod_usuarios_home')  # Redirige al home o la página deseada
        else:
            messages.error(request, 'Por favor, corrija los errores en el formulario.')
    else:
        form = UsuarioCreateForm()
    
    return render(request, 'usuarios/mod_personal/registrar_usuario.html', {'form': form})

# MOD CLIENTES
@login_required  
def mod_clientes_home(request):
    grupo_clientes = Cliente.objects.all()
    usuarios = grupo_clientes.filter(is_active=True)

    # Búsqueda por término
    search_term = request.GET.get('search', '')
    if search_term:
        usuarios = usuarios.filter(
            Q(nombre__icontains=search_term) |
            Q(apellido__icontains=search_term) |
            Q(CI__icontains=search_term)
        )

    # Paginación
    paginator = Paginator(usuarios, 8)  # Mostrar 8 clientes por página
    page_number = request.GET.get('page')
    usuarios_page = paginator.get_page(page_number)

    return render(request, 'usuarios/mod_cliente/mod_clientes_home.html', {
        'usuarios': usuarios_page,
        'search_term': search_term
    })

@login_required
def ver_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    return render(request, 'usuarios/mod_cliente/ver_cliente.html', {'cliente': cliente})

@login_required
def editar_cliente(request, cliente_id):
    usuario = get_object_or_404(Cliente, id=cliente_id)
    if request.method == 'POST':
        form = ClienteForm(request.POST, request.FILES, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente editado exitosamente.')
            return redirect('mod_clientes_home')
        else:
            # Imprimir errores en la consola para depuración
            print(form.errors)
            messages.error(request, 'Por favor, corrige los errores del formulario.')
    else:
        form = ClienteForm(instance=usuario)

    return render(request, 'usuarios/mod_cliente/editar_cliente.html', {'form': form, 'usuario': usuario})

@login_required
def eliminar_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)

    # Cambiar el estado del usuario
    if cliente.is_active:
        cliente.is_active = False  # Desactivar el usuario (Dado de baja)
        estado = "Dado de baja"
    else:
        cliente.is_active = True   # Activar el usuario
        estado = "Activo"

    cliente.save()  # Guardar los cambios en el estado

    # Mostrar un mensaje de éxito
    messages.success(request, f'Cliente {estado} exitosamente.')

    # Redirigir a la vista de la lista de usuarios
    return redirect('mod_clientes_home')

@login_required
def registrar_cliente(request):
    if request.method == 'POST':
        form = ClienteCreateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()  # Esto ya maneja la asociación del grupo en el método save()
            messages.success(request, 'Cliente creado exitosamente.')
            return redirect('mod_clientes_home')  # Redirige al home o la página deseada
        else:
            messages.error(request, 'Por favor, corrija los errores en el formulario.')
    else:
        form = ClienteCreateForm()
    
    return render(request, 'usuarios/mod_cliente/registrar_cliente.html', {'form': form})
#MOD ROLES
@login_required
def roles_list(request):
    roles = Group.objects.all().order_by('id')
    return render(request, 'usuarios/mod_personal/roles_list.html', {'roles': roles})
# Agregar Rol
def agregar_rol(request):
    if request.method == 'POST':
        nombre_rol = request.POST.get('nombre')
        if nombre_rol:
            if not Group.objects.filter(name=nombre_rol).exists():
                Group.objects.create(name=nombre_rol)
                messages.success(request, 'Rol creado exitosamente.')
            else:
                messages.error(request, 'Ya existe un rol con ese nombre.')
            return redirect('roles_list')
        else:
            messages.error(request, 'El nombre del rol no puede estar vacío.')
    return render(request, 'usuarios/mod_personal/agregar_rol.html')

# Editar Rol
def editar_rol(request, rol_id):
    rol = get_object_or_404(Group, id=rol_id)
    if request.method == 'POST':
        nuevo_nombre = request.POST.get('nombre')
        if nuevo_nombre:
            if not Group.objects.filter(name=nuevo_nombre).exclude(id=rol_id).exists():
                rol.name = nuevo_nombre
                rol.save()
                messages.success(request, 'Rol actualizado exitosamente.')
            else:
                messages.error(request, 'Ya existe un rol con ese nombre.')
            return redirect('roles_list')
        else:
            messages.error(request, 'El nombre del rol no puede estar vacío.')
    return render(request, 'usuarios/mod_personal/editar_rol.html', {'rol': rol})

#MOD VER PERFIL
@login_required
def ver_perfil(request):
    user = request.user  # Obtiene el usuario actualmente autenticado
    return render(request, 'usuarios/ver_perfil.html', {'user': user})
@login_required
def editar_perfil(request):
    if request.method == 'POST':
        form = UsuarioUpdateForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "¡Perfil actualizado correctamente!")
            return redirect('ver_perfil')
    else:
        form = UsuarioUpdateForm(instance=request.user)

    return render(request, 'usuarios/editar_perfil.html', {'form': form})

@login_required
def restablecer_contrasena(request):
    """Vista para restablecer la contraseña del usuario."""
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)  # Mantener sesión activa
            messages.success(request, "¡Contraseña actualizada correctamente!")
            return redirect('ver_perfil')
        else:
            messages.error(request, "Por favor corrige los errores abajo.")
    else:
        form = PasswordChangeForm(user=request.user)

    return render(request, 'usuarios/restablecer_contraseña.html', {'form': form})

#MOD PRODUCTOS
@login_required
def mod_productos_home(request):
    search_term = request.GET.get('search', '')
    productos = ProductoInventario.objects.filter(nombre__icontains=search_term, is_active=True)
    
    # Paginación
    paginator = Paginator(productos, 8)  # Muestra 10 productos por página
    page_number = request.GET.get('page')  # Número de página desde la URL
    productos_paginados = paginator.get_page(page_number)
    
    return render(request, 'productos/mod_productos.html', {
        'productos': productos_paginados,
        'search_term': search_term,
    })

@login_required
def ver_producto(request, producto_id):
    producto = get_object_or_404(ProductoInventario, id_producto=producto_id)
    return render(request, 'productos/ver_producto.html', {'producto': producto})

@login_required
def editar_producto(request, producto_id):
    producto = get_object_or_404(ProductoInventario, id_producto=producto_id)

    if request.method == 'POST':
        form = ProductoInventarioForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto actualizado exitosamente.')
            return redirect('mod_productos_home')
        else:
            # Imprimir errores en la consola para depuración
            print(form.errors)
            messages.error(request, 'Por favor, corrige los errores del formulario.')
    else:
        form = ProductoInventarioForm(instance=producto)

    return render(request, 'productos/editar_producto.html', {'form': form, 'producto': producto})

@login_required
def eliminar_producto(request, producto_id):
    producto = get_object_or_404(ProductoInventario, id_producto=producto_id)
    # Cambiar el estado
    if producto.is_active:
        producto.is_active = False  # Desactivar el usuario (Dado de baja)
        estado = "Dado de baja"
    else:
        producto.is_active = True   # Activar el usuario
        estado = "Activo"
    producto.save()  # Guardar los cambios en el estado
    # Mostrar un mensaje de éxito
    messages.success(request, f'Producto {estado} exitosamente.')
    # Redirigir a la vista de la lista de usuarios
    return redirect('mod_productos_home')

@login_required
def crear_producto(request):
    if request.method == 'POST':
        form = ProductoInventarioCreateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto creado exitosamente.')
            return redirect('mod_productos_home')  # Redirigir al panel de productos después de crear
        else:
            messages.error(request, 'Por favor, corrija los errores en el formulario.')
    else:
        form = ProductoInventarioCreateForm()
    
    return render(request, 'productos/crear_productos.html', {'form': form})

#MOD CATEGORíAs

@login_required
def mod_categorias_home(request):
    search_term = request.GET.get('search', '')
    categoria = Categoria.objects.filter(is_active=True)
    # buscar por el nombre de la categoria
    categoria = categoria.filter(nombre__icontains=search_term)
    
    categorias_inactivas = Categoria.objects.filter(is_active=False)
    
    
    # Paginación
    paginator = Paginator(categoria, 10)  # Muestra 10 productos por página
    page_number = request.GET.get('page')  # Número de página desde la URL
    categorias_paginadas = paginator.get_page(page_number)
    
    return render(request, 'categorias/mod_categorias.html', {
        'categoria': categorias_paginadas,
        'search_term': search_term,
    })

@login_required
def editar_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id_categoria=categoria_id)

    if request.method == 'POST':
        form = CategoriaForm(request.POST, request.FILES, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría actualizada exitosamente.')
            return redirect('mod_categorias_home')
        else:
            # Imprimir errores en la consola para depuración
            print(form.errors)
            messages.error(request, 'Por favor, corrige los errores del formulario.')
    else:
        form = CategoriaForm(instance=categoria)

    return render(request, 'categorias/editar_categoria.html', {'form': form, 'categoria': categoria})

@login_required
def eliminar_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id_categoria=categoria_id)
    # Cambiar el estado del usuario
    if categoria.is_active:
        categoria.is_active = False  # Desactivar el usuario (Dado de baja)
        estado = "Dado de baja"
    else:
        categoria.is_active = True   # Activar el usuario
        estado = "Activo"
    categoria.save()  # Guardar los cambios en el estado
    # Mostrar un mensaje de éxito
    messages.success(request, f'Categoría {estado} exitosamente.')
    # Redirigir a la vista de la lista de usuarios
    return redirect('mod_categorias_home')

@login_required
def crear_categoria(request):
    if request.method == 'POST':
        form = CategoriaCreateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría creada exitosamente.')
            return redirect('mod_categorias_home')  # Redirigir al panel de productos después de crear
        else:
            messages.error(request, 'Por favor, corrija los errores en el formulario.')
    else:
        form = CategoriaCreateForm()
    
    return render(request, 'categorias/crear_categorias.html', {'form': form})

#MOD VENTAS
@login_required  
def mod_ventas_home(request):
    # Obtener las ventas
    # ventas = Ventas.objects.all()
    #fecha en formato descendente
    ventas = Ventas.objects.filter(is_active=True).order_by('-fecha_registro')
    search_term = request.GET.get('search', '')
    if search_term:
        ventas = ventas.filter(
            
            Q(id_venta__icontains=search_term) |
            Q(id_cliente__nombre__icontains=search_term) |
            Q(id_cliente__apellido__icontains=search_term)
        )

    # Paginación
    paginator = Paginator(ventas, 8)  # Mostrar 8 clientes por página
    page_number = request.GET.get('page')
    ventas_page = paginator.get_page(page_number)

    return render(request, 'ventas/mod_ventas_home.html', {
        'ventas': ventas_page,
        'search_term': search_term
    })

@login_required
def ver_productos_venta(request, venta_id):
    # Obtener la venta y los productos relacionados
    venta = get_object_or_404(Ventas, id_venta=venta_id)
    productos = ProductosVenta.objects.filter(venta=venta)  # Suponiendo que DetalleVenta enlaza venta y productos
    
    context = {
        'venta': venta,
        'productos': productos,
    }
    return render(request, 'ventas/ver_productos_venta.html', context)

@login_required
def ver_detalle_venta(request, venta_id):
    # Obtener la venta y los productos asociados
    venta = get_object_or_404(Ventas, id_venta=venta_id)
    productos = ProductosVenta.objects.filter(venta=venta)
    
    context = {
        'venta': venta,
        'productos': productos,
    }
    return render(request, 'ventas/ver_detalle_venta.html', context)

@login_required
def editar_venta(request, venta_id):
    # Obtener la venta existente
    venta = get_object_or_404(Ventas, id_venta=venta_id)
    ItemFormSet = modelformset_factory(
        ProductosVenta, form=ProductosVentaForm, extra=1, can_delete=True
    )
    
    if request.method == "POST":
        venta_form = VentasForm(request.POST, instance=venta)
        items_formset = ItemFormSet(
            request.POST,
            queryset=ProductosVenta.objects.filter(venta=venta)
        )
        
        if venta_form.is_valid() and items_formset.is_valid():
            # Primero, manejar eliminaciones
            for form in items_formset.deleted_forms:
                if form.instance.pk:  # Asegurarse de que el objeto exista en la BD
                    form.instance.delete()
                else:
                    form.save(commit=False).delete()

            # Luego, procesar los formularios restantes
            for form in items_formset:
                if form.cleaned_data.get('DELETE', False):  # Omite eliminados
                    continue

                producto_venta = form.save(commit=False)
                producto_venta.venta = venta

                # Validar que el producto exista y esté asociado correctamente
                producto = form.cleaned_data.get('producto')
                if producto:
                    try:
                        producto_venta.producto = ProductoInventario.objects.get(
                            id_producto=producto.id_producto
                        )
                    except ProductoInventario.DoesNotExist:
                        form.add_error('producto', 'El producto relacionado ya no existe.')
                        continue
                else:
                    form.add_error('producto', 'El campo producto es obligatorio.')
                    continue

                # Guardar producto de la venta
                producto_venta.save()

            # Guardar la venta después de procesar los productos
            venta_form.save()

            # Redirigir al editar venta
            return redirect('editar_venta', venta_id=venta.id_venta)
            
    else:
        # Si es un GET, inicializa los formularios
        venta_form = VentasForm(instance=venta)
        items_formset = ItemFormSet(
            queryset=ProductosVenta.objects.filter(venta=venta)
        )

    # Renderizar el template con los formularios
    return render(request, 'ventas/editar_venta.html', {
        'venta_form': venta_form,
        'items_formset': items_formset,
        'costo_total': venta.calcular_costo_total()
    })

@login_required
def dar_de_baja_venta(request, venta_id):
    venta = get_object_or_404(Ventas, id_venta=venta_id)
    if venta.is_active == True:
        venta.is_active = False
        estado = "Dado de baja"
    else:
        venta.is_active = True
        estado = "Activo"
    venta.save()
    messages.success(request, f'Venta {estado} exitosamente.')
    return redirect('mod_ventas_home')

@login_required
#  ventas = Ventas.objects.filter(is_active=True)
def imprimir_venta(request, venta_id):
    ventas = get_object_or_404(Ventas, id_venta=venta_id)
    productos = ProductosVenta.objects.filter(venta=ventas) # Obtener los productos de la venta

    return render(request, 'ventas/imprimir_venta.html', {
        'ventas': ventas,
        'productos': productos,
    })

@login_required
@transaction.atomic
def crear_venta(request):
    if request.method == 'POST':
        venta_form = VentasForm(request.POST)
        if venta_form.is_valid():
            # Crear la venta
            venta = venta_form.save(commit=False)
            venta.id_vendedor = request.user
            venta.total = 0
            venta.save()

            productos_data = []
            for key, value in request.POST.items():
                if 'productos-' in key and '-producto_id' in key:
                    index = key.split('-')[1]  # Asegúrate de que el índice es único
                    producto_id = request.POST.get(f'productos-{index}-producto_id', '').strip()
                    cantidad = request.POST.get(f'productos-{index}-cantidad', '').strip()
                    # Validar que producto_id y cantidad no estén vacíos y sean números
                    if not producto_id.isdigit() or not cantidad.isdigit():
                        continue  # Ignorar filas con datos inválidos

                    productos_data.append({
                        'producto_id': int(producto_id),
                        'cantidad': int(cantidad),
                    })

            # Procesar productos y calcular el total
            total_venta = 0
            for producto_data in productos_data:
                producto_id = producto_data['producto_id']
                cantidad = producto_data['cantidad']

                try:
                    producto = ProductoInventario.objects.get(id_producto=producto_id)
                    subtotal = producto.precio_unitario * cantidad
                    total_venta += subtotal

                    # Crear el registro de ProductosVenta
                    ProductosVenta.objects.create(
                        venta=venta,
                        producto=producto,
                        cantidad=cantidad,
                        subtotal=subtotal,
                    )
                    producto.save()
                except ProductoInventario.DoesNotExist:
                    return JsonResponse({'error': f"Producto con ID {producto_id} no existe"}, status=400)

            # Actualizar el total de la venta
            venta.total = total_venta
            venta.save()

            return redirect('mod_ventas_home')
        else:
            return JsonResponse({'success': False, 'errors': venta_form.errors}, status=400)
    else:
        productos = ProductoInventario.objects.all()
        venta_form = VentasForm()
        return render(request, 'ventas/crear_venta.html', {'form': venta_form, 'productos': productos})

#MOD PEDIDOS
@login_required  
def mod_pedidos_home(request):
    pedidos = Pedidos.objects.filter(is_active=True).order_by('-fecha_registro')
    search_term = request.GET.get('search', '')
    if search_term:
        pedidos = pedidos.filter(
            
            Q(id_pedido__icontains=search_term) |
            Q(id_cliente__CI__icontains=search_term) |
            Q(id_cliente__nombre__icontains=search_term) |
            Q(id_cliente__apellido__icontains=search_term)
        )

    # Paginación
    paginator = Paginator(pedidos, 8)  # Mostrar 8 clientes por página
    page_number = request.GET.get('page')
    pedidos_page = paginator.get_page(page_number)

    return render(request, 'pedidos/mod_pedidos_home.html', {
        'pedidos': pedidos_page,
        'search_term': search_term
    })
   
@login_required 
def ver_productos_pedido(request, pedido_id):
    # Obtener la pedido y los productos relacionados
    pedido = get_object_or_404(Pedidos, id_pedido=pedido_id)
    productos = ProductosPedido.objects.filter(pedido=pedido)  # Suponiendo que DetalleVenta enlaza venta y productos
    
    context = {
        'pedido': pedido,
        'productos': productos,
    }
    return render(request, 'pedidos/ver_productos_pedido.html', context)

@login_required
def ver_detalle_pedido(request, pedido_id):
    # Obtener el pedido y los productos asociados
    pedido = get_object_or_404(Pedidos, id_pedido=pedido_id)
    productos = ProductosPedido.objects.filter(pedido=pedido)
    monto_faltante = pedido.costo_total - pedido.monto_pagado
    
    context = {
        'pedido': pedido,
        'productos': productos,
        'monto_faltante': monto_faltante
    }
    return render(request, 'pedidos/ver_detalle_pedido.html', context)

@login_required
def dar_de_baja_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedidos, id_pedido=pedido_id)
    if pedido.is_active == True:
        pedido.is_active = False
        estado = "Dado de baja"
    else:
        pedido.is_active = True
        estado = "Activo"
    pedido.save()
    messages.success(request, f'Pedido {estado} exitosamente.')
    return redirect('mod_pedidos_home')

@login_required
def imprimir_pedido(request, pedido_id):
    pedidos = get_object_or_404(Pedidos, id_pedido=pedido_id)
    productos = ProductosPedido.objects.filter(pedido=pedidos) # Obtener los productos de la venta
    monto_faltante = pedidos.costo_total - pedidos.monto_pagado

    return render(request, 'pedidos/imprimir_pedido.html', {
        'pedidos': pedidos,
        'productos': productos,
        'monto_faltante': monto_faltante
    })
    
@login_required
@transaction.atomic
def crear_pedido(request):
    if request.method == 'POST':
        venta_form = PedidosForm(request.POST)
        if venta_form.is_valid():
            # Crear la venta
            venta = venta_form.save(commit=False)
            venta.id_vendedor = request.user
            venta.total = 0
            venta.save()

            productos_data = []
            for key, value in request.POST.items():
                if 'productos-' in key and '-producto_id' in key:
                    index = key.split('-')[1]  # Asegúrate de que el índice es único
                    producto_id = request.POST.get(f'productos-{index}-producto_id', '').strip()
                    cantidad = request.POST.get(f'productos-{index}-cantidad', '').strip()
                    # Validar que producto_id y cantidad no estén vacíos y sean números
                    if not producto_id.isdigit() or not cantidad.isdigit():
                        continue  # Ignorar filas con datos inválidos

                    productos_data.append({
                        'producto_id': int(producto_id),
                        'cantidad': int(cantidad),
                    })

            # Procesar productos y calcular el total
            total_venta = 0
            for producto_data in productos_data:
                producto_id = producto_data['producto_id']
                cantidad = producto_data['cantidad']

                try:
                    producto = ProductoInventario.objects.get(id_producto=producto_id)
                    subtotal = producto.precio_unitario * cantidad
                    total_venta += subtotal

                    # Crear el registro de ProductosVenta
                    ProductosPedido.objects.create(
                        pedido=venta,
                        producto=producto,
                        cantidad=cantidad,
                        subtotal=subtotal,
                    )

                    # Actualizar el stock del producto
                    # producto.cantidad_stock -= cantidad
                    producto.save()
                except ProductoInventario.DoesNotExist:
                    return JsonResponse({'error': f"Producto con ID {producto_id} no existe"}, status=400)

            # Actualizar el total de la venta
            venta.total = total_venta
            venta.save()

            return redirect('mod_pedidos_home')
        else:
            return JsonResponse({'success': False, 'errors': venta_form.errors}, status=400)
    else:
        productos = ProductoInventario.objects.all()
        venta_form = PedidosForm()
        return render(request, 'pedidos/crear_pedido.html', {'form': venta_form, 'productos': productos})
    
#MOD REPORTES
@login_required
def mod_reportes_home(request):
    # Fecha y hora actual en la zona horaria local
    fecha_actual = localtime(now())

    # Determinar el primer y último día del mes actual
    primer_dia_mes = fecha_actual.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if fecha_actual.month == 12:
        primer_dia_mes_siguiente = fecha_actual.replace(year=fecha_actual.year + 1, month=1, day=1)
    else:
        primer_dia_mes_siguiente = fecha_actual.replace(month=fecha_actual.month + 1, day=1)

    # Total ventas del mes usando rango de fechas
    total_ventas_mes = Ventas.objects.filter(
        fecha_registro__gte=primer_dia_mes,
        fecha_registro__lt=primer_dia_mes_siguiente
    ).aggregate(total=Sum('costo_total'))['total'] or 0

    # Total clientes
    total_clientes = Cliente.objects.count()

    # Total productos
    total_productos = ProductoInventario.objects.count()

    # Pedidos pendientes
    pedidos_pendientes = Pedidos.objects.filter(estado="Pendiente").count()

    context = {
        'fecha_actual': fecha_actual,
        'total_ventas_mes': total_ventas_mes,
        'total_clientes': total_clientes,
        'total_productos': total_productos,
        'pedidos_pendientes': pedidos_pendientes,
    }

    return render(request, 'reportes/mod_reportes_home.html', context)

@login_required
def reporte_clientes(request):
    clientes_qs = Cliente.objects.all()
    ventas_qs = Ventas.objects.all()

    # --- Filtros GET ---
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    cliente_id = request.GET.get('cliente')
    estado = request.GET.get('estado')
    export_format = request.GET.get('export')

    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, "%Y-%m-%d")
            ventas_qs = ventas_qs.filter(fecha_registro__gte=fecha_desde_dt)
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, "%Y-%m-%d")
            ventas_qs = ventas_qs.filter(fecha_registro__lte=fecha_hasta_dt)
        except ValueError:
            pass
    if cliente_id:
        ventas_qs = ventas_qs.filter(id_cliente_id=cliente_id)
    if estado:
        ventas_qs = ventas_qs.filter(estado__iexact=estado)

    # --- Agrupamos ventas por cliente ---
    clientes_data = (
        ventas_qs.values("id_cliente__id", "id_cliente__nombre", "id_cliente__apellido")
        .annotate(
            total_ventas=Count("id_venta"),
            total_ingresos=Sum("costo_total")
        )
        .order_by("-total_ingresos")
    )

    # --- Exportar a Excel ---
    if export_format == "excel":
        import pandas as pd
        data = []
        for c in clientes_data:
            data.append({
                "Cliente": f"{c['id_cliente__nombre']} {c['id_cliente__apellido']}",
                "Ventas Realizadas": c["total_ventas"],
                "Total Comprado (Bs)": float(c["total_ingresos"] or 0),
            })
        df = pd.DataFrame(data)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Clientes")
            ws = writer.sheets['Clientes']
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="BDD7EE")
                cell.font = Font(bold=True)
        output.seek(0)
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Reporte_Clientes.xlsx'
        return response

    # --- Exportar a PDF ---
    if export_format == "pdf":
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        title = Paragraph("Reporte de Clientes", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))

        data = [["Cliente", "Ventas Realizadas", "Total Comprado (Bs)"]]
        for c in clientes_data:
            data.append([
                f"{c['id_cliente__nombre']} {c['id_cliente__apellido']}",
                str(c["total_ventas"]),
                f"{c['total_ingresos'] or 0:.2f}"
            ])
        table = Table(data, repeatRows=1, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F81BD")),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('FONTSIZE', (0,1), (-1,-1), 10),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.lightgrey])
        ]))
        elements.append(table)
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Reporte_Clientes.pdf"'
        response.write(pdf)
        return response

    # --- Estadísticas ---
    total_clientes = clientes_data.count()
    cliente_top = clientes_data.first() if clientes_data else None

    context = {
        "clientes": clientes_qs,
        "clientes_data": clientes_data,
        "total_clientes": total_clientes,
        "cliente_top": cliente_top,
    }
    return render(request, 'reportes/clientes.html', context)

#Para Excel

@login_required
def reporte_ventas(request):
    ventas_qs = Ventas.objects.all()
    clientes = Cliente.objects.all()

    # Filtros GET
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    cliente_id = request.GET.get('cliente')
    estado = request.GET.get('estado')
    export_format = request.GET.get('export')  # 'excel' o 'pdf'

    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, "%Y-%m-%d")
            ventas_qs = ventas_qs.filter(fecha_registro__gte=fecha_desde_dt)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            # Añadimos 1 día al final para incluir todo el día hasta las 23:59
            fecha_hasta_dt = datetime.strptime(fecha_hasta, "%Y-%m-%d")
            ventas_qs = ventas_qs.filter(fecha_registro__lte=fecha_hasta_dt)
        except ValueError:
            pass

    if cliente_id:
        ventas_qs = ventas_qs.filter(id_cliente=cliente_id)

    if estado:
        ventas_qs = ventas_qs.filter(estado__iexact=estado)
    # --- Filtrado igual que en reporte_ventas ---
    ventas_qs = Ventas.objects.all()
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    cliente_id = request.GET.get('cliente')
    estado = request.GET.get('estado')

    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, "%Y-%m-%d")
            ventas_qs = ventas_qs.filter(fecha_registro__gte=fecha_desde_dt)
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, "%Y-%m-%d")
            ventas_qs = ventas_qs.filter(fecha_registro__lte=fecha_hasta_dt)
        except ValueError:
            pass
    if cliente_id:
        ventas_qs = ventas_qs.filter(id_cliente=cliente_id)
    if estado:
        ventas_qs = ventas_qs.filter(estado__iexact=estado)

    # --- Exportar a Excel ---
    
    if export_format == "excel":
        import pandas as pd
        data = []
        for venta in ventas_qs:
            data.append({
                "ID": venta.id_venta,
                "Fecha": venta.fecha_registro.strftime("%d/%m/%Y %H:%M"),
                "Cliente": venta.id_cliente.nombre,
                "Total Bs": float(venta.costo_total),
                "Estado": venta.estado,
            })
        df = pd.DataFrame(data)
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Ventas")
            # Bordes: aplicar formato de tabla
            ws = writer.sheets['Ventas']
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
            
            # Aplicar borde y alineación a todas las celdas
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            # Opcional: fondo para encabezado
            for cell in ws[1]:
                    cell.fill = PatternFill("solid", fgColor="BDD7EE")
                    cell.font = Font(bold=True)
        output.seek(0)
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Reporte_Ventas.xlsx'
        return response

    # --- Exportar a PDF con tabla y bordes ---
    if export_format == "pdf":
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        styles = getSampleStyleSheet()
        title = Paragraph("Reporte de Ventas", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))

        # Datos de la tabla
        data = [["ID", "Fecha", "Cliente", "Total Bs", "Estado"]]
        for venta in ventas_qs:
            data.append([
                str(venta.id_venta),
                venta.fecha_registro.strftime("%d/%m/%Y %H:%M"),
                venta.id_cliente.nombre,
                f"{venta.costo_total:.2f}",
                venta.estado
            ])

        # Crear tabla con estilo
        table = Table(data, repeatRows=1, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F81BD")),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('FONTSIZE', (0,1), (-1,-1), 10),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.lightgrey])
        ]))

        elements.append(table)
        doc.build(elements)

        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Reporte_Ventas.pdf"'
        response.write(pdf)
        return response
    # Estadísticas
    total_ventas = ventas_qs.count()
    total_ingresos = ventas_qs.aggregate(total=Sum('costo_total'))['total'] or 0
    promedio_venta = total_ingresos / total_ventas if total_ventas > 0 else 0
    productos_vendidos = (
        ProductosVenta.objects.filter(venta__in=ventas_qs)
        .aggregate(total=Sum("cantidad"))["total"] or 0
    )
    
    # --- Gráfico de tendencia de ventas ---
    ventas_por_fecha = (
        ventas_qs.values("fecha_registro")
        .annotate(total=Sum("costo_total"))
        .order_by("fecha_registro")
    )
    fechas_grafico = [v["fecha_registro"].strftime("%d/%m/%Y") for v in ventas_por_fecha if v["fecha_registro"]]
    ventas_grafico = [float(v["total"] or 0) for v in ventas_por_fecha]
    
    print("VENTAS POR FECHA:", list(ventas_por_fecha))
    print("FECHAS GRAFICO:", fechas_grafico)
    print("VENTAS GRAFICO:", ventas_grafico)
     # --- Top productos más vendidos ---
    productos_qs = (
        ProductosVenta.objects.filter(venta__in=ventas_qs)
        .values("producto__nombre")
        .annotate(total_vendido=Sum("cantidad"))
        .order_by("-total_vendido")[:5]
    )
    productos_labels = [p["producto__nombre"] for p in productos_qs]
    productos_data = [int(p["total_vendido"] or 0) for p in productos_qs]
    # Paginación
    paginator = Paginator(ventas_qs.order_by('-fecha_registro'), 100)
    page_number = request.GET.get('page')
    ventas_page = paginator.get_page(page_number)

    context = {
        "ventas": ventas_page,
        "clientes": clientes,
        "total_ventas": total_ventas,
        "total_ingresos": total_ingresos,
        "promedio_venta": promedio_venta,
        "productos_vendidos": productos_vendidos,
        "fechas_grafico": fechas_grafico,
        "ventas_grafico": ventas_grafico,
        "productos_labels": productos_labels,
        "productos_data": productos_data,
    }
    return render(request, 'reportes/ventas.html', context)

@login_required
def reporte_productos(request):
    productos_qs = ProductoInventario.objects.select_related('id_categoria').all()

    # Filtros GET
    categoria_id = request.GET.get('categoria')
    estado_stock = request.GET.get('estado_stock')
    producto_buscar = request.GET.get('producto')
    export_format = request.GET.get('export')  # 'excel' o 'pdf'

    if categoria_id:
        productos_qs = productos_qs.filter(id_categoria_id=categoria_id)
    if producto_buscar:
        productos_qs = productos_qs.filter(nombre__icontains=producto_buscar)

    # Calcular valor_total y estado_stock sin modificar el modelo
    productos_list = []
    for p in productos_qs:
        valor_total = p.precio_unitario * p.cantidad_stock
        if p.cantidad_stock == 0:
            estado = "Agotado"
        elif p.cantidad_stock <= 10:
            estado = "Bajo"
        elif p.cantidad_stock <= 50:
            estado = "Medio"
        else:
            estado = "Alto"
        productos_list.append({
            "id": p.id_producto,
            "nombre": p.nombre,
            "categoria": p.id_categoria.nombre,
            "stock": p.cantidad_stock,
            "precio_unitario": p.precio_unitario,
            "valor_total": valor_total,
            "estado": estado,
            "fecha_actualizacion": p.fecha_registro
        })

    # Filtrar por estado_stock si aplica
    if estado_stock:
        productos_list = [p for p in productos_list if p['estado'].lower() == estado_stock.lower()]

    # --- Exportar a Excel ---
    if export_format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos"

        # Encabezados
        headers = ["ID", "Producto", "Categoría", "Stock", "Precio Unitario", "Valor Total", "Estado", "Última Actualización"]
        ws.append(headers)

        for p in productos_list:
            ws.append([
                p["id"],
                p["nombre"],
                p["categoria"],
                p["stock"],
                float(p["precio_unitario"]),
                float(p["valor_total"]),
                p["estado"],
                p["fecha_actualizacion"].strftime("%d/%m/%Y %H:%M")
            ])

        # Formato de tabla
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="BDD7EE")
            cell.font = Font(bold=True)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Reporte_Productos.xlsx'
        return response

    # --- Exportar a PDF ---
    if export_format == "pdf":
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        styles = getSampleStyleSheet()
        title = Paragraph("Reporte de Inventario", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))

        data = [["ID", "Producto", "Categoría", "Stock", "Precio ud.", "Valor Total", "Estado", "Registro"]]
        for p in productos_list:
            data.append([
                str(p["id"]),
                p["nombre"],
                p["categoria"],
                str(p["stock"]),
                f"{p['precio_unitario']:.2f}",
                f"{p['valor_total']:.2f}",
                p["estado"],
                p["fecha_actualizacion"].strftime("%d/%m/%Y %H:%M")
            ])

        table = Table(data, repeatRows=1, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F81BD")),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('FONTSIZE', (0,1), (-1,-1), 10),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.lightgrey])
        ]))

        elements.append(table)
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Reporte_Productos.pdf"'
        response.write(pdf)
        return response

    # Estadísticas
    total_productos = len(productos_list)
    valor_inventario = sum([p["valor_total"] for p in productos_list])
    productos_bajo_stock = sum([1 for p in productos_list if p["estado"] == "Bajo"])
    productos_agotados = sum([1 for p in productos_list if p["estado"] == "Agotado"])

    # Datos para gráficos
    categorias = Categoria.objects.all()
    categorias_labels = [c.nombre for c in categorias]
    categorias_data = [len([p for p in productos_list if p["categoria"] == c.nombre]) for c in categorias]

    stock_normal = len([p for p in productos_list if p["estado"] in ["Alto", "Medio"]])
    stock_bajo = len([p for p in productos_list if p["estado"] == "Bajo"])
    stock_agotado = len([p for p in productos_list if p["estado"] == "Agotado"])

    # Paginación
    paginator = Paginator(productos_list, 10)
    page_number = request.GET.get('page')
    productos_page = paginator.get_page(page_number)

    context = {
        "productos": productos_page,
        "categorias": categorias,
        "total_productos": total_productos,
        "valor_inventario": valor_inventario,
        "productos_bajo_stock": productos_bajo_stock,
        "productos_agotados": productos_agotados,
        "categorias_labels": categorias_labels,
        "categorias_data": categorias_data,
        "stock_normal": stock_normal,
        "stock_bajo": stock_bajo,
        "stock_agotado": stock_agotado,
        "request": request
    }

    return render(request, "reportes/productos.html", context)

@login_required
def reporte_financiero(request):
    return render(request, 'reportes/financiero.html')

@login_required
def dashboard(request):
    # --- Filtros GET (igual que reporte_ventas) ---
    ventas_qs = Ventas.objects.all()
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    cliente_id = request.GET.get('cliente')
    estado = request.GET.get('estado')

    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, "%Y-%m-%d")
            ventas_qs = ventas_qs.filter(fecha_registro__gte=fecha_desde_dt)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, "%Y-%m-%d")
            ventas_qs = ventas_qs.filter(fecha_registro__lte=fecha_hasta_dt)
        except ValueError:
            pass

    if cliente_id:
        ventas_qs = ventas_qs.filter(id_cliente=cliente_id)

    if estado:
        ventas_qs = ventas_qs.filter(estado=estado.upper())

    # --- KPIs ---
    total_ingresos = ventas_qs.aggregate(total=Sum('costo_total'))['total'] or 0
    total_ventas = ventas_qs.count()
    productos_vendidos = ProductosVenta.objects.filter(venta__in=ventas_qs).aggregate(total=Sum('cantidad'))['total'] or 0
    clientes_activos = Cliente.objects.filter(ventas_como_cliente__estado='COMPLETADO').distinct().count()
    productos_stock = ProductoInventario.objects.filter(cantidad_stock__gt=0).count()

    # --- Gráfico de Tendencia de Ventas ---
    ventas_por_fecha = (
        ventas_qs
        .annotate(fecha=TruncDate('fecha_registro'))
        .values('fecha_registro')
        .annotate(total=Sum('costo_total'))
        .order_by('fecha_registro')
    )

    fechas_grafico = [v['fecha_registro'].strftime("%d/%m/%Y") for v in ventas_por_fecha if v['fecha_registro']]
    ventas_grafico = [float(v['total'] or 0) for v in ventas_por_fecha]
    
    print("VENTAS POR FECHA:", list(ventas_por_fecha))
    print("FECHAS GRAFICO:", fechas_grafico)
    print("VENTAS GRAFICO:", ventas_grafico)


    # --- Distribución por Categoría ---
    categorias_qs = (
        ProductosVenta.objects.filter(venta__in=ventas_qs)
        .values('producto__id_categoria__nombre')
        .annotate(total=Sum('cantidad'))
        .order_by('-total')
    )
    categorias_labels = [c['producto__id_categoria__nombre'] for c in categorias_qs]
    categorias_data = [int(c['total'] or 0) for c in categorias_qs]

    # --- Resumen por Módulos ---
    resumen_modulos = {
        'usuarios': {
            'activos': Usuario.objects.filter(is_active=True).count(),
            'total': Usuario.objects.count()
        },
        'productos': {
            'activos': ProductoInventario.objects.filter(is_active=True).count(),
            'total': ProductoInventario.objects.count()
        },
        'clientes': {
            'activos': clientes_activos,
            'total': Cliente.objects.count()
        },
        'pedidos': {
            'activos': Pedidos.objects.filter(is_active=True).count(),
            'total': Pedidos.objects.count()
        }
    }

    # --- Alertas importantes ---
    alertas = {
        'stock_bajo': ProductoInventario.objects.filter(cantidad_stock__lt=10).count(),
        'pedidos_pendientes': Pedidos.objects.filter(estado='Pendiente').count(),
    }

    # --- Top productos ---
    top_productos_qs = (
        ProductosVenta.objects.values('producto__nombre')
        .annotate(total_vendido=Sum('cantidad'))
        .order_by('-total_vendido')[:5]
    )

    # --- Actividad reciente ---
    actividad_reciente = ventas_qs.order_by('-fecha_registro')[:5]

    # --- Clientes para filtro ---
    clientes = Cliente.objects.all()

    context = {
        'total_ingresos': total_ingresos,
        'total_ventas': total_ventas,
        'productos_vendidos': productos_vendidos,
        'clientes_activos': clientes_activos,
        'productos_stock': productos_stock,
        'fechas_grafico': fechas_grafico,
        'ventas_grafico': ventas_grafico,
        'categorias_labels': categorias_labels,
        'categorias_data': categorias_data,
        'resumen_modulos': resumen_modulos,
        'alertas': alertas,
        'top_productos': top_productos_qs,
        'actividad_reciente': actividad_reciente,
        'clientes': clientes,
        'filtros': {
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'cliente_id': int(cliente_id) if cliente_id else '',
            'estado': estado or ''
        }
    }

    return render(request, 'reportes/dashboard.html', context)

@login_required
def reporte_pedidos(request):
    pedidos_qs = Pedidos.objects.all()
    clientes = Cliente.objects.all()
    vendedores = Usuario.objects.all()

    # Filtros GET
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    cliente_id = request.GET.get('cliente')
    vendedor_id = request.GET.get('vendedor')
    estado = request.GET.get('estado')
    export_format = request.GET.get('export')  # 'excel' o 'pdf'

    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, "%Y-%m-%d")
            pedidos_qs = pedidos_qs.filter(fecha_registro__gte=fecha_desde_dt)
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, "%Y-%m-%d")
            pedidos_qs = pedidos_qs.filter(fecha_registro__lte=fecha_hasta_dt)
        except ValueError:
            pass
    if cliente_id:
        pedidos_qs = pedidos_qs.filter(id_cliente_id=cliente_id)
    if vendedor_id:
        pedidos_qs = pedidos_qs.filter(id_vendedor_id=vendedor_id)
    if estado:
        pedidos_qs = pedidos_qs.filter(estado__iexact=estado)

    # --- Exportar a Excel ---
    if export_format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Pedidos"
        # Encabezado
        headers = ["ID Pedido", "Fecha", "Cliente", "Vendedor", "Beneficiario", "Monto Pagado", "Total", "Estado"]
        ws.append(headers)

        # Datos
        for pedido in pedidos_qs:
            ws.append([
                pedido.id_pedido,
                pedido.fecha_registro.strftime("%d/%m/%Y %H:%M"),
                pedido.id_cliente.nombre,
                pedido.id_vendedor.username,
                pedido.beneficiario,
                float(pedido.monto_pagado),
                float(pedido.costo_total),
                pedido.estado,
            ])

        # Formato de tabla
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        # Fondo encabezado
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="BDD7EE")
            cell.font = Font(bold=True)

        # Respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Reporte_Pedidos.xlsx'
        return response

    # --- Exportar a PDF ---
    if export_format == "pdf":
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        styles = getSampleStyleSheet()
        title = Paragraph("Reporte de Pedidos", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))

        data = [["ID Pedido", "Fecha", "Cliente", "Vendedor", "Beneficiario", "Monto Pagado", "Total", "Estado"]]
        for pedido in pedidos_qs:
            data.append([
                str(pedido.id_pedido),
                pedido.fecha_registro.strftime("%d/%m/%Y %H:%M"),
                pedido.id_cliente.nombre,
                pedido.id_vendedor.username,
                pedido.beneficiario,
                f"{pedido.monto_pagado:.2f}",
                f"{pedido.costo_total:.2f}",
                pedido.estado
            ])

        table = Table(data, repeatRows=1, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F81BD")),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('FONTSIZE', (0,1), (-1,-1), 10),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.lightgrey])
        ]))
        elements.append(table)
        doc.build(elements)

        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Reporte_Pedidos.pdf"'
        response.write(pdf)
        return response

    # --- Estadísticas ---
    total_pedidos = pedidos_qs.count()
    total_monto = pedidos_qs.aggregate(total=Sum('monto_pagado'))['total'] or 0
    total_completados = pedidos_qs.filter(estado='COMPLETADO').count()
    total_pendientes = pedidos_qs.filter(estado='PENDIENTE').count()
    total_cancelados = pedidos_qs.filter(estado='CANCELADO').count()

    # Gráficos
    pedidos_por_fecha = pedidos_qs.values("fecha_registro").annotate(total=Sum("costo_total")).order_by("fecha_registro")
    fechas_grafico = [p["fecha_registro"].strftime("%d/%m/%Y") for p in pedidos_por_fecha if p["fecha_registro"]]
    pedidos_grafico = [float(p["total"] or 0) for p in pedidos_por_fecha]

    productos_qs = ProductosPedido.objects.filter(pedido__in=pedidos_qs).values("producto__nombre").annotate(total_pedido=Sum("cantidad")).order_by("-total_pedido")[:5]
    productos_labels = [p["producto__nombre"] for p in productos_qs]
    productos_data = [int(p["total_pedido"] or 0) for p in productos_qs]

    # Paginación
    paginator = Paginator(pedidos_qs.order_by('-fecha_registro'), 100)
    page_number = request.GET.get('page')
    pedidos_page = paginator.get_page(page_number)

    context = {
        "pedidos": pedidos_page,
        "clientes": clientes,
        "vendedores": vendedores,
        "total_pedidos": total_pedidos,
        "total_monto": total_monto,
        "total_completados": total_completados,
        "total_pendientes": total_pendientes,
        "total_cancelados": total_cancelados,
        "fechas_grafico": fechas_grafico,
        "pedidos_grafico": pedidos_grafico,
        "productos_labels": productos_labels,
        "productos_data": productos_data,
        "request": request
    }

    return render(request, 'reportes/pedidos.html', context)